import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from core.logger import log

class ExcelBuilder:
    def __init__(self, json_path, output_path="output_report.xlsx"):
        self.json_path = json_path
        self.output_path = output_path
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Report"
        self.current_row = 1
        
        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self.header_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def load_data(self):
        if not os.path.exists(self.json_path):
            raise FileNotFoundError(f"❌ Could not find {self.json_path}. Please create it first.")
        with open(self.json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data.get("document", {})

    def get_max_columns(self, document):
        max_cols = 1
        for table in document.get("tables", []):
            if "headers" in table:
                max_cols = max(max_cols, len(table["headers"]))
        return max_cols

    def write_merged_text(self, text, max_cols, is_bold, font_size, alignment):
        """Writes titles/footers and dynamically scales the row height for massive paragraphs."""
        if not text:
            return
        cell = self.ws.cell(row=self.current_row, column=1, value=text)
        self.ws.merge_cells(start_row=self.current_row, start_column=1, end_row=self.current_row, end_column=max_cols)
        
        cell.font = Font(name='Nirmala UI', bold=is_bold, size=font_size)
        cell.alignment = alignment
        
        # 🚀 THE NEW SMART HEIGHT CALCULATOR FOR MERGED FOOTERS/NOTES
        # Estimate that each column can hold roughly 15 characters safely
        chars_per_line = max(max_cols * 15, 30) 
        estimated_lines = str(text).count('\n') + (len(str(text)) // chars_per_line) + 1
        
        # Scale the height based on the number of lines and the font size
        self.ws.row_dimensions[self.current_row].height = estimated_lines * (font_size * 1.5)
        self.current_row += 1

    def _autofit_columns(self):
        merged_cells_map = set()
        for merged_range in self.ws.merged_cells.ranges:
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_cells_map.add((row, col))

        max_col_idx = self.ws.max_column
        max_row_idx = self.ws.max_row
        
        for col_idx in range(1, max_col_idx + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            for row_idx in range(1, max_row_idx + 1):
                if (row_idx, col_idx) in merged_cells_map:
                    continue
                
                cell = self.ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    # Estimate the width needed for the longest single line in this cell
                    lines = str(cell.value).split('\n')
                    longest_line = max([len(line) for line in lines])
                    if longest_line > max_length:
                        max_length = longest_line
            
            # Constraints: Min width 12, Max 45
            adjusted_width = max(12, min(max_length + 4, 45))
            self.ws.column_dimensions[column_letter].width = adjusted_width

    def build(self):
        log.info("🚀 Booting Smart Excel Builder...")
        document = self.load_data()
        if not document:
            log.error("❌ Invalid JSON format.")
            return

        max_cols = self.get_max_columns(document)
        log.info(f"📊 Detected maximum document width: {max_cols} columns")

        main_title = document.get("main_title", {})
        self.write_merged_text(
            main_title.get("text", ""), max_cols, 
            main_title.get("is_bold", True), main_title.get("font_size", 14), 
            self.center_align
        )

        for subtitle in document.get("subtitles", []):
            self.write_merged_text(
                subtitle.get("text", ""), max_cols, 
                subtitle.get("is_bold", True), subtitle.get("font_size", 12), 
                self.center_align
            )
            
        self.current_row += 1 

        for table in document.get("tables", []):
            table_title = table.get("table_title", "")
            if table_title:
                self.write_merged_text(table_title, max_cols, True, 12, self.left_align)

            headers = table.get("headers", [])
            for col_idx, header in enumerate(headers, start=1):
                cell = self.ws.cell(row=self.current_row, column=col_idx, value=header.get("column_name", ""))
                cell.font = Font(name='Nirmala UI', bold=header.get("is_bold", True), size=11)
                cell.alignment = self.center_align
                cell.border = self.thin_border
                cell.fill = self.header_fill
            self.current_row += 1

            for row_data in table.get("rows", []):
                max_lines_in_row = 1
                for col_idx, value in enumerate(row_data, start=1):
                    cell = self.ws.cell(row=self.current_row, column=col_idx, value=str(value))
                    cell.font = Font(name='Nirmala UI', size=11)
                    cell.alignment = self.center_align
                    cell.border = self.thin_border
                    
                    # More robust row height estimator for table columns
                    lines = str(value).count('\n') + (len(str(value)) // 30) + 1
                    if lines > max_lines_in_row:
                        max_lines_in_row = lines
                        
                self.ws.row_dimensions[self.current_row].height = max_lines_in_row * 16
                self.current_row += 1
                
            self.current_row += 1 

        footer = document.get("footer", {})
        # If AI hallucinates and returns a list (e.g., ["footer text"]), convert to string
        if isinstance(footer, list):
            footer_text = "\n".join([str(i) for i in footer])
            footer = {"text": footer_text, "is_bold": False, "font_size": 11}
        # If AI hallucinates and returns a pure string, convert to dict
        elif isinstance(footer, str):
            footer = {"text": footer, "is_bold": False, "font_size": 11}
        self.write_merged_text(
            footer.get("text", ""), max_cols, 
            footer.get("is_bold", False), footer.get("font_size", 11), 
            self.left_align
        )

        self._autofit_columns()

        self.wb.save(self.output_path)
        log.info(f"✅ Success! Smart Auto-Fit Report saved to: {self.output_path}")