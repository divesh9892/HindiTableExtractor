import unittest
import json
import os
import tempfile
from openpyxl import load_workbook
from core.excel_builder import ExcelBuilder

class TestExcelBuilder(unittest.TestCase):

    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.test_json_path = os.path.join(self.temp_dir.name, "test_data.json")
        self.test_excel_path = os.path.join(self.temp_dir.name, "test_output.xlsx")

    def tearDown(self):
        self.temp_dir.cleanup()

    def create_mock_json(self, data):
        with open(self.test_json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f)

    def test_missing_json_file(self):
        """Proves the script gracefully catches missing files."""
        builder = ExcelBuilder(json_path="does_not_exist.json")
        with self.assertRaises(FileNotFoundError):
            builder.load_data()

    def test_standard_generation_and_autofit(self):
        """Proves the Excel file is generated and the Auto-Fit doesn't crash."""
        mock_data = {
            "document": {
                "main_title": {"text": "Test Report", "is_bold": True, "font_size": 14},
                "tables": [
                    {
                        "headers": [{"column_name": "ID"}, {"column_name": "Very Long Column Name To Test AutoFit"}],
                        "rows": [["1", "Short Data"], ["2", "Another piece of data"]]
                    }
                ]
            }
        }
        self.create_mock_json(mock_data)
        
        builder = ExcelBuilder(self.test_json_path, self.test_excel_path)
        builder.build()
        
        self.assertTrue(os.path.exists(self.test_excel_path))
        wb = load_workbook(self.test_excel_path)
        ws = wb.active
        
        self.assertEqual(ws["A1"].value, "Test Report")
        
        width_A = ws.column_dimensions['A'].width
        width_B = ws.column_dimensions['B'].width
        self.assertGreater(width_B, width_A)

    def test_missing_optional_fields(self):
        """Proves the script handles missing headers, titles, and rows safely."""
        mock_data = {
            "document": {
                "tables": [
                    {
                        "headers": [{"column_name": "Single Column"}],
                    }
                ]
            }
        }
        self.create_mock_json(mock_data)
        builder = ExcelBuilder(self.test_json_path, self.test_excel_path)
        builder.build()
        
        self.assertTrue(os.path.exists(self.test_excel_path))
        wb = load_workbook(self.test_excel_path)
        
        # 🚀 FIXED: Because there is no Title, Row 1 is skipped, and the header drops to Row 2!
        self.assertEqual(wb.active["A2"].value, "Single Column")

if __name__ == "__main__":
    unittest.main(verbosity=2)